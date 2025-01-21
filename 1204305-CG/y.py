from openpyxl import Workbook
from openpyxl.styles import Font


def calculate_points(x, y, xc, yc):
  return [
      (x, y),
      (xc + x, yc + y),
      (xc - x, yc + y),
      (xc - x, yc - y),
      (xc + x, yc - y)
  ]


# Initial coordinates
coordinates = [
(0, 20)
,(1, 20)
,(2, 20)
,(3, 19)
,(4, 19)
,(5, 18)
,(6, 17)
,(7, 16)
,(7, 15)
,(8, 14)
,(8, 13)
,(9, 12)
,(9, 11)
,(10, 10)
,(10, 9)
,(10, 8)
,(10, 7)
,(10, 6)
,(11, 5)
,(11, 4)
,(11, 3)
,(11, 2)
,(11, 1)
,(11, 0)
]

# Center point
xc, yc = 2, 7

dy = 20

# Create workbook and select active sheet
wb = Workbook()
ws = wb.active

# Write headers
headers = ['(x,y)', 'P1(xc+x,yc+y)', 'P2(xc-x,yc+y)', 'P3(xc-x,yc-y)', 'P4(xc+x,yc-y)']
for col, header in enumerate(headers, 1):
  cell = ws.cell(row=1, column=col)
  cell.value = header
  cell.font = Font(bold=True)

# Fill data
for row, coord in enumerate(coordinates, 2):
  points = calculate_points(coord[0], coord[1], xc, yc)
  formulas = [
    f"({coord[0]}, {coord[1]})",
    f"({points[1][0]}, {points[1][1]} - {dy} - 2) = ({points[1][0]}, {points[1][1] - dy - 2})",
    f"({points[2][0]}, {points[2][1]} - {dy} - 2) = ({points[2][0]}, {points[2][1] - dy - 2})",
    f"({points[3][0]}, {points[3][1]} + {dy} + 2) = ({points[3][0]}, {points[3][1] + dy + 2})",
    f"({points[4][0]}, {points[4][1]} + {dy} + 2) = ({points[4][0]}, {points[4][1] + dy + 2})"
  ]
  for col, formula in enumerate(formulas, 1):
    ws.cell(row=row, column=col, value=formula)

# Save the workbook
wb.save('coordinate_transformations2.xlsx')
